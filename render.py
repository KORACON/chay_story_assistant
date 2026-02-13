import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from PIL import Image
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


# -------------------------
# Пути
# -------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
FONTS_DIR = BASE_DIR / "fonts"
ASSETS_DIR = BASE_DIR / "assets"

ACCESS_FILE = DATA_DIR / "access.json"

ASSET_PRODUCTS_BG = ASSETS_DIR / "Фон.png"
ASSET_TEA_BANK_BG = ASSETS_DIR / "Ценник Банки 70мм x 70мм - 300ppi Фон.png"
ASSET_TEA_BOX_BG = ASSETS_DIR / "Ценник Коробки 160мм x 20мм - 300ppi Фон.png"
ASSET_TIPS_FRONT_BG = ASSETS_DIR / "Главная Чаевые Фон.png"
ASSET_TIPS_BACK_BG = ASSETS_DIR / "Бэк Чаевые.png"


# -------------------------
# Цвета
# -------------------------
ORANGE = (0xF6 / 255, 0x76 / 255, 0x3C / 255)  # #F6763C
CREAM = (0xF4 / 255, 0xEF / 255, 0xE8 / 255)   # #F4EFE8
LINE = (0xC1 / 255, 0xBA / 255, 0xB1 / 255)    # #C1BAB1

QR_BG_HEX = "FEF6E9"
QR_FG_HEX = "231F20"


# -------------------------
# Шрифты
# -------------------------
@dataclass
class Fonts:
    regular: str
    medium: str
    semibold: str
    bold: str


def register_unbounded_fonts() -> Fonts:
    bold_path = FONTS_DIR / "Unbounded-Bold.ttf"
    med_path = FONTS_DIR / "Unbounded-Medium.ttf"
    semi_path = FONTS_DIR / "Unbounded-SemiBold.ttf"
    reg_path = FONTS_DIR / "Unbounded-Regular.ttf"

    for p in [bold_path, med_path, semi_path]:
        if not p.exists():
            raise FileNotFoundError(f"Не найден шрифт: {p}. Положи файл в папку fonts/")

    if not reg_path.exists():
        reg_path = med_path  # fallback

    pdfmetrics.registerFont(TTFont("Unbounded-Bold", str(bold_path)))
    pdfmetrics.registerFont(TTFont("Unbounded-Medium", str(med_path)))
    pdfmetrics.registerFont(TTFont("Unbounded-SemiBold", str(semi_path)))
    pdfmetrics.registerFont(TTFont("Unbounded-Regular", str(reg_path)))

    return Fonts(
        regular="Unbounded-Regular",
        medium="Unbounded-Medium",
        semibold="Unbounded-SemiBold",
        bold="Unbounded-Bold",
    )


def ensure_assets_exist():
    required = [
        ASSET_PRODUCTS_BG,
        ASSET_TEA_BANK_BG,
        ASSET_TEA_BOX_BG,
        ASSET_TIPS_FRONT_BG,
        ASSET_TIPS_BACK_BG,
    ]
    for p in required:
        if not p.exists():
            raise FileNotFoundError(f"Не найден ассет: {p}. Положи файл в папку assets/.")


# -------------------------
# Утилиты
# -------------------------
def safe_filename(name: str, max_len: int = 80) -> str:
    name = (name or "").strip()
    name = re.sub(r"[\\/:*?\"<>|\n\r\t]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = "item"
    if len(name) > max_len:
        name = name[:max_len].rstrip()
    return name


def unique_names(base_names: List[str]) -> List[str]:
    used: Dict[str, int] = {}
    out = []
    for n in base_names:
        if n not in used:
            used[n] = 1
            out.append(n)
        else:
            used[n] += 1
            out.append(f"{n}_{used[n]}")
    return out


def parse_int_number(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    s = str(value).strip().replace(",", ".")
    if re.fullmatch(r"\d+(\.0+)?", s):
        return int(float(s))
    return None


def category_from_price(price: int) -> str:
    if price <= 20:
        return "A"
    if price <= 35:
        return "A+"
    if price <= 55:
        return "A++"
    return "ПРЕМИУМ"


def hours_word(n: int) -> str:
    n = abs(int(n))
    n100 = n % 100
    n10 = n % 10
    if 11 <= n100 <= 14:
        return "часов"
    if n10 == 1:
        return "час"
    if 2 <= n10 <= 4:
        return "часа"
    return "часов"


def normalize_sentence_case(text: str) -> str:
    """
    Делает:
    - первая буква заглавная
    - после . ! ? следующая буква заглавная
    """
    s = re.sub(r"\s+", " ", (text or "").strip())
    if not s:
        return s

    # Разбиваем на предложения, сохраняя разделители
    parts = re.split(r"([.!?]+)", s)
    out = []
    for i in range(0, len(parts), 2):
        chunk = parts[i].strip()
        punct = parts[i + 1] if i + 1 < len(parts) else ""

        if chunk:
            # найти первую букву
            m = re.search(r"[A-Za-zА-Яа-яЁё]", chunk)
            if m:
                idx = m.start()
                chunk = chunk[:idx] + chunk[idx].upper() + chunk[idx + 1:]
        out.append(chunk + punct)

    res = " ".join([x.strip() for x in out if x.strip()])
    res = re.sub(r"\s+([.!?])", r"\1", res)
    return res.strip()


def img_size(path: Path) -> Tuple[int, int]:
    with Image.open(path) as im:
        return im.size


def pdf_with_background(page_w: int, page_h: int, bg_path: Path) -> Tuple[io.BytesIO, canvas.Canvas]:
    buff = io.BytesIO()
    c = canvas.Canvas(buff, pagesize=(page_w, page_h))
    c.drawImage(ImageReader(str(bg_path)), 0, 0, page_w, page_h, mask="auto")
    return buff, c


def text_width(font_name: str, size: int, text: str) -> float:
    return pdfmetrics.stringWidth(text, font_name, size)


def break_long_word(word: str, font_name: str, size: int, max_width: float) -> List[str]:
    parts, buf = [], ""
    for ch in word:
        test = buf + ch
        if text_width(font_name, size, test) <= max_width:
            buf = test
        else:
            if buf:
                parts.append(buf)
                buf = ch
            else:
                parts.append(ch)
                buf = ""
    if buf:
        parts.append(buf)
    return parts


def wrap_lines(
    text: str,
    font_name: str,
    size: int,
    max_width: float,
    max_lines: int,
    allow_word_break: bool = True,
) -> Optional[List[str]]:
    """
    Разбивает текст на строки по словам.
    - allow_word_break=True: если встречается слово, которое не влезает, режем его по буквам (старое поведение).
    - allow_word_break=False: переносов по буквам НЕ делаем; если слово не влезает — возвращаем None (тогда уменьшаем шрифт).
    """
    text = re.sub(r"\s+", " ", (text or "").strip())
    if not text:
        return None

    words = text.split(" ")
    lines: List[str] = []
    current = ""

    def push(line: str):
        lines.append(line)

    i = 0
    while i < len(words):
        w = words[i]

        # слово само по себе не влезает
        if text_width(font_name, size, w) > max_width:
            if not allow_word_break:
                return None
            pieces = break_long_word(w, font_name, size, max_width)
            words = words[:i] + pieces + words[i + 1 :]
            w = words[i]

        test = (current + " " + w).strip() if current else w
        if text_width(font_name, size, test) <= max_width:
            current = test
            i += 1
        else:
            push(current)
            current = ""
            if len(lines) >= max_lines:
                return None

    if current:
        push(current)

    return lines if len(lines) <= max_lines else None


def fit_text(
    text: str,
    font_name: str,
    max_size: int,
    min_size: int,
    max_width: float,
    max_lines: int,
    allow_word_break: bool = True,
) -> Tuple[int, List[str]]:
    """Подбирает максимально возможный размер шрифта, чтобы текст влез."""
    for size in range(max_size, min_size - 1, -1):
        lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break)
        if lines:
            return size, lines

    size = min_size
    lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break) or [
        ((text or "").strip()[:20] + "…").strip()
    ]
    return size, lines
    size = min_size
    lines = wrap_lines(text, font_name, size, max_width, max_lines) or [((text or "").strip()[:20] + "…").strip()]
    return size, lines



def fit_text_in_box(
    text: str,
    font_name: str,
    max_size: int,
    min_size: int,
    max_width: float,
    max_lines: int,
    max_height: float,
    line_height: float,
    allow_word_break: bool = True,
) -> Tuple[int, List[str]]:
    """Как fit_text, но ещё учитывает высоту блока (max_height)."""
    for size in range(max_size, min_size - 1, -1):
        lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break)
        if not lines:
            continue
        needed_h = size * line_height * len(lines)
        if needed_h <= max_height:
            return size, lines

    size = min_size
    lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break) or [
        ((text or "").strip()[:20] + "…").strip()
    ]
    return size, lines

def draw_centered_multiline(
    c: canvas.Canvas,
    lines: List[str],
    font_name: str,
    font_size: int,
    center_x: float,
    center_y: float,
    color_rgb: Tuple[float, float, float],
    line_height: float = 1.2,
):
    c.setFillColorRGB(*color_rgb)
    c.setFont(font_name, font_size)

    block_h = font_size * line_height * len(lines)
    y = center_y + (block_h / 2) - font_size
    for line in lines:
        c.drawCentredString(center_x, y, line)
        y -= font_size * line_height


def draw_multiline_in_rect(
    c: canvas.Canvas,
    lines: List[str],
    font_name: str,
    font_size: int,
    x0: float,
    x1: float,
    center_y: float,
    color_rgb: Tuple[float, float, float],
    *,
    align: str = "center",
    line_height: float = 1.2,
):
    """Рисует многострочный текст внутри прямоугольника по X (x0..x1), по Y центрируется вокруг center_y.
    align: 'left' | 'center' | 'right'
    """
    c.setFillColorRGB(*color_rgb)
    c.setFont(font_name, font_size)

    block_h = font_size * line_height * len(lines)
    y = center_y + (block_h / 2) - font_size

    for line in lines:
        if align == "left":
            c.drawString(x0, y, line)
        elif align == "right":
            c.drawRightString(x1, y, line)
        else:
            c.drawCentredString((x0 + x1) / 2, y, line)
        y -= font_size * line_height


def fit_text_above_line(
    text: str,
    font_name: str,
    max_size: int,
    min_size: int,
    max_width: float,
    max_lines: int,
    *,
    y_top: float,
    y_line: float,
    clearance: float,
    line_height: float = 1.05,
    allow_word_break: bool = False,
) -> Tuple[int, List[str], float]:
    """Подбирает максимально крупный текст, который:
    - влезает по ширине/строкам
    - помещается в область от y_line+clearance до y_top
    - при этом низ глифов (baseline + descent) у нижней строки >= y_line+clearance

    Возвращает: (size, lines, last_baseline_y)
    """
    text = re.sub(r"\s+", " ", (text or "").strip())
    if not text:
        return min_size, [""], y_line + clearance

    y_min_glyph = y_line + clearance  # минимальная Y низа букв

    for size in range(max_size, min_size - 1, -1):
        lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break)
        if not lines:
            continue

        ascent = pdfmetrics.getAscent(font_name, size)
        descent = abs(pdfmetrics.getDescent(font_name, size))

        # baseline последней строки так, чтобы низ глифов был на y_min_glyph:
        last_baseline = y_min_glyph + descent

        # baseline первой строки
        first_baseline = last_baseline + (len(lines) - 1) * (size * line_height)

        # верх глифов первой строки
        top_glyph = first_baseline + ascent

        if top_glyph <= y_top:
            return size, lines, last_baseline

    # fallback
    size = min_size
    lines = wrap_lines(text, font_name, size, max_width, max_lines, allow_word_break=allow_word_break) or [text]
    ascent = pdfmetrics.getAscent(font_name, size)
    descent = abs(pdfmetrics.getDescent(font_name, size))
    last_baseline = (y_line + clearance) + descent
    first_baseline = last_baseline + (len(lines) - 1) * (size * line_height)
    top_glyph = first_baseline + ascent
    # если всё равно не влезло, просто сдвинем вверх
    if top_glyph > y_top:
        shift = top_glyph - y_top
        last_baseline -= shift
    return size, lines, last_baseline


def draw_multiline_above_line(
    c: canvas.Canvas,
    lines: List[str],
    font_name: str,
    font_size: int,
    center_x: float,
    last_baseline_y: float,
    color_rgb: Tuple[float, float, float],
    *,
    line_height: float = 1.05,
):
    """Рисует строки так, чтобы baseline последней строки был last_baseline_y (якорь снизу)."""
    c.setFillColorRGB(*color_rgb)
    c.setFont(font_name, font_size)

    first_baseline = last_baseline_y + (len(lines) - 1) * (font_size * line_height)
    y = first_baseline
    for line in lines:
        c.drawCentredString(center_x, y, line)
        y -= font_size * line_height

def draw_brand_ci(c: canvas.Canvas, fonts: Fonts, page_w: int, y: float, size: int):
    # Ч(оранж) + АЙНАЯ(светл) + пробел + И(оранж) + СТОРИЯ(светл)
    font = fonts.medium
    seg1, seg2, seg3, seg4, seg5 = "Ч", "АЙНАЯ", " ", "И", "СТОРИЯ"

    w1 = text_width(font, size, seg1)
    w2 = text_width(font, size, seg2)
    w3 = text_width(font, size, seg3)
    w4 = text_width(font, size, seg4)
    w5 = text_width(font, size, seg5)
    total = w1 + w2 + w3 + w4 + w5

    x = (page_w - total) / 2

    c.setFont(font, size)
    c.setFillColorRGB(*ORANGE)
    c.drawString(x, y, seg1); x += w1

    c.setFillColorRGB(*CREAM)
    c.drawString(x, y, seg2); x += w2
    c.drawString(x, y, seg3); x += w3

    c.setFillColorRGB(*ORANGE)
    c.drawString(x, y, seg4); x += w4

    c.setFillColorRGB(*CREAM)
    c.drawString(x, y, seg5)


def format_price(price: int) -> str:
    return f"{price}₽"  # без пробела


# -------------------------
# PDF генерация
# -------------------------
def make_pdf_products_two_sides(fonts: Fonts, name: str, price: int, hours: int) -> bytes:
    # нормализация названия товара
    name = normalize_sentence_case(name)

    w, h = img_size(ASSET_PRODUCTS_BG)
    buff, c = pdf_with_background(w, h, ASSET_PRODUCTS_BG)

    # FRONT
    # Бренд должен быть на том же уровне и того же размера, что и на обратной стороне.
    # Ориентируемся на BACK (y=585, size=36).
    draw_brand_ci(c, fonts, w, y=585, size=36)

    size_name, lines_name = fit_text(name, fonts.bold, max_size=72, min_size=34, max_width=w - 140, max_lines=2)
    draw_centered_multiline(c, lines_name, fonts.bold, size_name, w / 2, 355, CREAM, line_height=1.15)

    # Цена почти внизу
    price_text = format_price(price)
    size_price, lines_price = fit_text(price_text, fonts.semibold, max_size=60, min_size=28, max_width=w - 200, max_lines=1)
    draw_centered_multiline(c, lines_price, fonts.semibold, size_price, w / 2, 80, ORANGE, line_height=1.0)

    c.showPage()

    # BACK
    c.drawImage(ImageReader(str(ASSET_PRODUCTS_BG)), 0, 0, w, h, mask="auto")
    draw_brand_ci(c, fonts, w, y=585, size=36)

    title = "Срок хранения"
    size_t, lines_t = fit_text(title, fonts.bold, max_size=72, min_size=36, max_width=w - 160, max_lines=2)
    draw_centered_multiline(c, lines_t, fonts.bold, size_t, w / 2, 355, CREAM, line_height=1.05)

    phrase = f"{hours} {hours_word(hours)}"
    size_h, lines_h = fit_text(phrase, fonts.semibold, max_size=60, min_size=28, max_width=w - 200, max_lines=1)
    draw_centered_multiline(c, lines_h, fonts.semibold, size_h, w / 2, 80, ORANGE, line_height=1.0)

    c.save()
    return buff.getvalue()


def make_pdf_tea_bank(fonts: Fonts, tea_type: str, name: str, price: int) -> bytes:
    w, h = img_size(ASSET_TEA_BANK_BG)
    buff, c = pdf_with_background(w, h, ASSET_TEA_BANK_BG)

    # ---------- Бренд (ЧАЙНАЯ ИСТОРИЯ) ВВЕРХУ, увеличен
    draw_brand_ci(c, fonts, w, y=1505, size=70)

    # ---------- Категория (1 строка, максимально крупно)
    cat = category_from_price(price)
    size_cat, lines_cat = fit_text_in_box(
        cat,
        fonts.medium,
        max_size=85,
        min_size=18,
        max_width=w - 220,
        max_lines=1,
        max_height=120,
        line_height=1.0,
        allow_word_break=False,
    )
    draw_centered_multiline(c, lines_cat, fonts.medium, size_cat, w / 2, 210, CREAM, line_height=1.0)

    # ---------- Тип чая (якорим над линией, чтобы НИКОГДА не залезал на белую полоску)
    y_top = 1425
    y_line = 855
    clearance = 36  # держим низ букв выше линии (примерно как на синей отметке)
    size_tt, lines_tt, last_baseline = fit_text_above_line(
        tea_type,
        fonts.bold,
        max_size=180,
        min_size=34,
        max_width=w - 240,
        max_lines=2,
        y_top=y_top,
        y_line=y_line,
        clearance=clearance,
        line_height=1.05,
        allow_word_break=False,
    )
    draw_multiline_above_line(c, lines_tt, fonts.bold, size_tt, w / 2, last_baseline, ORANGE, line_height=1.05)

    # верхняя линия (должна быть белой)
    c.setStrokeColorRGB(*LINE)
    c.setLineWidth(6)
    c.setLineCap(1)
    c.line(190, 855, w - 190, 855)

    # ---------- Название (между линиями)
    top_y2 = 840
    bottom_y2 = 615  # чуть выше нижней линии (600)
    box_h2 = top_y2 - bottom_y2
    center_y2 = (top_y2 + bottom_y2) / 2

    size_nm, lines_nm = fit_text_in_box(
        name,
        fonts.medium,
        max_size=124,
        min_size=22,
        max_width=w - 260,
        max_lines=2,
        max_height=box_h2,
        line_height=1.2,
        allow_word_break=False,
    )
    draw_centered_multiline(c, lines_nm, fonts.medium, size_nm, w / 2, center_y2, CREAM, line_height=1.2)

    # нижняя линия чуть короче
    c.setStrokeColorRGB(*LINE)
    c.setLineWidth(6)
    c.setLineCap(1)
    c.line(260, 600, w - 260, 600)

    # ---------- Цена (крупнее)
    price_text = format_price(price)
    size_p, lines_p = fit_text(
        price_text,
        fonts.bold,
        max_size=110,
        min_size=30,
        max_width=w - 340,
        max_lines=1,
        allow_word_break=False,
    )
    draw_centered_multiline(c, lines_p, fonts.bold, size_p, w / 2, 470, ORANGE, line_height=1.0)

    # ---------- Бренд (крупнее)
    c.save()
    return buff.getvalue()



def make_pdf_tea_box(fonts: Fonts, tea_type: str, name: str, price: int) -> bytes:
    w, h = img_size(ASSET_TEA_BOX_BG)
    buff, c = pdf_with_background(w, h, ASSET_TEA_BOX_BG)

    x1, x2 = 457, 1524
    c.setStrokeColorRGB(*LINE)
    c.setLineWidth(6)
    c.setLineCap(1)
    c.line(x1, 30, x1, h - 30)
    c.line(x2, 30, x2, h - 30)

    pad = 60
    left_x0, left_x1 = pad, x1 - pad
    mid_x0, mid_x1 = x1 + pad, x2 - pad
    right_x0, right_x1 = x2 + pad, w - pad

    size_tt, lines_tt = fit_text(
        tea_type,
        fonts.bold,
        max_size=72,
        min_size=28,
        max_width=(left_x1 - left_x0),
        max_lines=2,
        allow_word_break=False,
    )
    draw_multiline_in_rect(c, lines_tt, fonts.bold, size_tt, left_x0, left_x1, h / 2, ORANGE, align="left", line_height=1.0)

    size_nm, lines_nm = fit_text(name, fonts.medium, max_size=60, min_size=24, max_width=(mid_x1 - mid_x0), max_lines=2)
    draw_centered_multiline(c, lines_nm, fonts.medium, size_nm, (mid_x0 + mid_x1) / 2, h / 2 + 3, CREAM, line_height=1.2)

    price_text = format_price(price)
    size_p, lines_p = fit_text(price_text, fonts.bold, max_size=60, min_size=24, max_width=(right_x1 - right_x0), max_lines=1)
    draw_multiline_in_rect(c, lines_p, fonts.bold, size_p, right_x0, right_x1, 150, ORANGE, align="right", line_height=1.0)

    cat = category_from_price(price)
    size_c, lines_c = fit_text(cat, fonts.medium, max_size=36, min_size=18, max_width=(right_x1 - right_x0), max_lines=1)
    draw_multiline_in_rect(c, lines_c, fonts.medium, size_c, right_x0, right_x1, 80, CREAM, align="right", line_height=1.0)

    c.save()
    return buff.getvalue()


def make_styled_qr_png(data: str, size_px: int = 180) -> bytes:
    import qrcode
    from qrcode.constants import ERROR_CORRECT_H

    try:
        from qrcode.image.styledpil import StyledPilImage
        from qrcode.image.styles.moduledrawers import RoundedModuleDrawer
        from qrcode.image.styles.colormasks import SolidFillColorMask
        from qrcode.image.styles.eyedrawers import RoundedEyeDrawer

        qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_H, box_size=10, border=1)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(
            image_factory=StyledPilImage,
            module_drawer=RoundedModuleDrawer(),
            eye_drawer=RoundedEyeDrawer(),
            color_mask=SolidFillColorMask(back_color=f"#{QR_BG_HEX}", front_color=f"#{QR_FG_HEX}"),
        ).convert("RGBA")
    except Exception:
        qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_H, box_size=10, border=1)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color=f"#{QR_FG_HEX}", back_color=f"#{QR_BG_HEX}").convert("RGBA")

    img = img.resize((size_px, size_px), Image.NEAREST)
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def make_pdf_tips_two_sides(fonts: Fonts, person_name: str, goal: str, link: str) -> bytes:
    w, h = img_size(ASSET_TIPS_FRONT_BG)
    buff, c = pdf_with_background(w, h, ASSET_TIPS_FRONT_BG)

    # Имя: максимально крупно, без переносов по буквам
    size_n, lines_n = fit_text(
        person_name,
        fonts.bold,
        max_size=86,
        min_size=28,
        max_width=w - 120,
        max_lines=2,
        allow_word_break=False,
    )

    # Цель: перенос по словам, если не влезает — уменьшаем шрифт (без переносов по буквам)
    size_g, lines_g = fit_text(
        goal,
        fonts.regular,
        max_size=48,
        min_size=16,
        max_width=w - 140,
        max_lines=3,
        allow_word_break=False,
    )

    # Вертикальная раскладка: имя и цель не должны пересекаться
    name_lh = 1.05
    goal_lh = 1.15
    # Отступ между именем и целью: небольшой как в шаблоне + адаптивно сжимается на длинных текстах
    base_gap = 20
    gap = base_gap
    if len(lines_n) >= 2 or len(lines_g) >= 2:
        gap = 14
    if len(lines_n) >= 2 and len(lines_g) >= 2:
        gap = 10

    name_h = size_n * name_lh * len(lines_n)
    goal_h = size_g * goal_lh * len(lines_g)

    group_center_y = 505  # визуальный центр зоны текста
    name_center_y = group_center_y + (goal_h + gap) / 2
    goal_center_y = group_center_y - (name_h + gap) / 2

    # Не даём цели опуститься на QR (верх QR-блока около 311)
    safe_goal_bottom = 330
    goal_bottom = goal_center_y - goal_h / 2
    if goal_bottom < safe_goal_bottom:
        shift = safe_goal_bottom - goal_bottom
        name_center_y += shift
        goal_center_y += shift

    draw_centered_multiline(c, lines_n, fonts.bold, size_n, w / 2, name_center_y, CREAM, line_height=name_lh)
    draw_centered_multiline(c, lines_g, fonts.regular, size_g, w / 2, goal_center_y, ORANGE, line_height=goal_lh)

    # QR
    box_left = 200
    box_bottom = 62
    box_size = 249
    qr_size = 180
    qr_left = int(box_left + (box_size - qr_size) / 2)
    qr_bottom = int(box_bottom + (box_size - qr_size) / 2)

    qr_png = make_styled_qr_png(link, size_px=qr_size)
    c.drawImage(ImageReader(io.BytesIO(qr_png)), qr_left, qr_bottom, qr_size, qr_size, mask="auto")

    c.showPage()
    c.drawImage(ImageReader(str(ASSET_TIPS_BACK_BG)), 0, 0, w, h, mask="auto")
    c.save()
    return buff.getvalue()



# -------------------------
# Excel шаблоны и чтение
# -------------------------
def build_xlsx_tea_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Чай"
    headers = ["Тип чая", "Наименование", "Цена (число)"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16

    for _ in range(2, 202):
        ws.append(["", "", ""])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_xlsx_products_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = ["Название", "Цена (число)", "Срок хранения (часы, число)"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28

    for _ in range(2, 202):
        ws.append(["", "", ""])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def load_rows_tea(xlsx_bytes: bytes) -> List[Tuple[str, str, int]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        tea_type, name, price = (r[0], r[1], r[2]) if len(r) >= 3 else (None, None, None)
        if (tea_type is None or str(tea_type).strip() == "") and (name is None or str(name).strip() == "") and (price is None or str(price).strip() == ""):
            continue

        tea_type_s = str(tea_type).strip() if tea_type is not None else ""
        name_s = str(name).strip() if name is not None else ""
        price_i = parse_int_number(price)

        if not tea_type_s:
            raise ValueError("В Excel для чая найдено пустое поле «Тип чая».")
        if not name_s:
            raise ValueError("В Excel для чая найдено пустое поле «Наименование».")
        if price_i is None or price_i < 0 or price_i > 1_000_000:
            raise ValueError("В Excel для чая «Цена» должна быть числом (0…1000000).")

        rows.append((tea_type_s, name_s, price_i))

    if not rows:
        raise ValueError("Excel пустой: заполни хотя бы одну строку.")
    return rows


def load_rows_products(xlsx_bytes: bytes) -> List[Tuple[str, int, int]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name, price, hours = (r[0], r[1], r[2]) if len(r) >= 3 else (None, None, None)
        if (name is None or str(name).strip() == "") and (price is None or str(price).strip() == "") and (hours is None or str(hours).strip() == ""):
            continue

        name_s = str(name).strip() if name is not None else ""
        price_i = parse_int_number(price)
        hours_i = parse_int_number(hours)

        if not name_s:
            raise ValueError("В Excel для товаров найдено пустое поле «Название».")
        if price_i is None or price_i < 0 or price_i > 1_000_000:
            raise ValueError("В Excel для товаров «Цена» должна быть числом (0…1000000).")
        if hours_i is None or hours_i < 0 or hours_i > 24 * 365:
            raise ValueError("В Excel для товаров «Срок хранения» должен быть числом часов (0…8760).")

        rows.append((name_s, price_i, hours_i))

    if not rows:
        raise ValueError("Excel пустой: заполни хотя бы одну строку.")
    return rows
